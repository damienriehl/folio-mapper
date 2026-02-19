from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from app.models.parse_models import ParseItem, ParseResult
from app.services.hierarchy_detector import (
    detect_hierarchy,
    parse_hierarchical,
)


ALLOWED_EXTENSIONS = {".xlsx", ".csv", ".tsv", ".txt", ".md"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 50_000
MAX_COLUMNS = 50


def _detect_headers(rows: list[list[str]]) -> tuple[list[str] | None, list[list[str]]]:
    """Heuristic: if the first row has all non-empty cells and the second row
    has a different pattern, treat the first row as headers."""
    if len(rows) < 2:
        return None, rows

    first = rows[0]
    second = rows[1]

    first_filled = all(c.strip() for c in first)
    second_has_blanks = any(not c.strip() for c in second)
    multi_col = len(first) >= 2

    if first_filled and second_has_blanks:
        # Hierarchical pattern: header row then indented data
        return [c.strip() for c in first], rows[1:]

    if first_filled and multi_col and len(rows) >= 3:
        # Multi-column flat: check if remaining rows share a consistent column count
        # and first row is plausibly a header (heuristic: all cells are short strings)
        data_rows = rows[1:]
        first_all_short = all(len(c.strip()) < 50 for c in first)
        if first_all_short:
            return [c.strip() for c in first], data_rows

    return None, rows


def _parse_tabular(
    rows: list[list[str]], filename: str | None = None
) -> ParseResult:
    """Process tabular data: detect headers, check for hierarchy, extract items."""
    if not rows:
        return ParseResult(
            format="flat",
            items=[],
            total_items=0,
            source_filename=filename,
        )

    headers, data_rows = _detect_headers(rows)

    if detect_hierarchy(data_rows):
        return parse_hierarchical(data_rows, headers=headers, filename=filename)

    # Flat: extract items from first non-empty column of each row
    items: list[ParseItem] = []
    for i, row in enumerate(data_rows):
        text = ""
        for cell in row:
            if cell.strip():
                text = cell.strip()
                break
        if text:
            items.append(ParseItem(text=text, index=len(items)))

    raw_preview = rows[:5]

    return ParseResult(
        format="flat",
        items=items,
        total_items=len(items),
        headers=headers,
        source_filename=filename,
        raw_preview=raw_preview,
    )


def _read_csv(content: bytes, delimiter: str = ",") -> list[list[str]]:
    """Read CSV/TSV content with encoding fallback."""
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = content.decode(encoding)
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            return [row for row in reader]
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode file with supported encodings")


def _read_excel(content: bytes) -> list[list[str]]:
    """Read the first sheet of an Excel file."""
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return rows


def _read_text(content: bytes) -> list[list[str]]:
    """Read a plain text file as single-column rows."""
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = content.decode(encoding)
            return [[line] for line in text.splitlines() if line.strip()]
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode file with supported encodings")


def parse_file(content: bytes, filename: str) -> ParseResult:
    """Dispatch file parsing based on extension."""
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type: {ext}. "
            f"Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )

    if len(content) > MAX_FILE_SIZE:
        raise ValueError(f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB.")

    if ext == ".xlsx":
        rows = _read_excel(content)
    elif ext == ".csv":
        rows = _read_csv(content, delimiter=",")
    elif ext == ".tsv":
        rows = _read_csv(content, delimiter="\t")
    elif ext in (".txt", ".md"):
        rows = _read_text(content)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    # Enforce row and column limits to prevent memory exhaustion
    if len(rows) > MAX_ROWS:
        raise ValueError(
            f"Too many rows ({len(rows):,}). Maximum: {MAX_ROWS:,}."
        )
    if rows and len(rows[0]) > MAX_COLUMNS:
        raise ValueError(
            f"Too many columns ({len(rows[0])}). Maximum: {MAX_COLUMNS}."
        )

    result = _parse_tabular(rows, filename=filename)
    return result
