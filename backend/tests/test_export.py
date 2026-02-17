"""Tests for the export service and router."""

from __future__ import annotations

import csv
import io
import json
from unittest.mock import MagicMock, patch

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from app.main import app
from app.models.export_models import (
    ExportConcept,
    ExportOptions,
    ExportRequest,
    ExportRow,
)
from app.services.export_scope import enrich_concept, expand_scope
from app.services.export_service import (
    generate_csv,
    generate_excel,
    generate_html,
    generate_json,
    generate_json_ld,
    generate_markdown,
    generate_rdf_turtle,
)


@pytest.fixture
def anyio_backend():
    return "asyncio"


@pytest.fixture
async def client():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c


def _make_concept(**overrides) -> ExportConcept:
    defaults = {
        "label": "Criminal Law",
        "iri": "https://folio.openlegalstandard.org/RCriminalLaw",
        "iri_hash": "RCriminalLaw",
        "branch": "Area of Law",
        "score": 85.0,
        "definition": "Law dealing with crimes and punishments",
        "translations": {"es": "Derecho Penal", "fr": "Droit pénal"},
    }
    defaults.update(overrides)
    return ExportConcept(**defaults)


def _make_row(**overrides) -> ExportRow:
    defaults = {
        "item_index": 0,
        "source_text": "DUI Defense",
        "ancestry": ["Criminal Law"],
        "selected_concepts": [_make_concept()],
        "note": "Reviewed",
        "status": "completed",
    }
    defaults.update(overrides)
    return ExportRow(**defaults)


def _make_options(**overrides) -> ExportOptions:
    defaults = {
        "format": "csv",
        "include_confidence": True,
        "include_notes": True,
        "include_reasoning": False,
        "iri_format": "hash",
        "languages": [],
        "include_hierarchy": True,
    }
    defaults.update(overrides)
    return ExportOptions(**defaults)


def _make_request(**overrides) -> ExportRequest:
    defaults = {
        "rows": [_make_row()],
        "options": _make_options(),
        "source_file": "test.txt",
        "session_created": None,
    }
    defaults.update(overrides)
    return ExportRequest(**defaults)


# --- CSV tests ---


def test_csv_produces_valid_csv_with_bom():
    req = _make_request()
    data = generate_csv(req)
    # UTF-8 BOM
    assert data[:3] == b"\xef\xbb\xbf"
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    assert rows[0] == ["Source", "Label", "IRI", "Branch", "Confidence", "Notes", "Definition", "Alt Labels", "Examples", "Hierarchy"]
    assert rows[1][0] == "DUI Defense"
    assert rows[1][1] == "Criminal Law"
    assert rows[1][2] == "RCriminalLaw"


def test_csv_includes_language_columns():
    req = _make_request(options=_make_options(languages=["es", "fr"]))
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    assert "Label (Español)" in rows[0]
    assert "Label (Français)" in rows[0]
    # Check translation values
    es_idx = rows[0].index("Label (Español)")
    assert rows[1][es_idx] == "Derecho Penal"


def test_csv_empty_selections():
    row = _make_row(selected_concepts=[], note="Skipped")
    req = _make_request(rows=[row])
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    assert len(rows) == 2  # header + 1 row
    assert rows[1][0] == "DUI Defense"
    assert rows[1][1] == ""  # empty label


# --- Excel tests ---


def test_excel_produces_valid_workbook():
    req = _make_request()
    data = generate_excel(req)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "FOLIO Mappings"
    assert ws.cell(1, 1).value == "Source"
    assert ws.cell(1, 1).font.bold is True
    assert ws.cell(2, 1).value == "DUI Defense"
    assert ws.cell(2, 2).value == "Criminal Law"


# --- JSON tests ---


def test_json_produces_valid_structure():
    req = _make_request()
    data = generate_json(req)
    parsed = json.loads(data)
    assert "exported" in parsed
    assert parsed["source_file"] == "test.txt"
    assert parsed["total_items"] == 1
    assert len(parsed["mappings"]) == 1
    mapping = parsed["mappings"][0]
    assert mapping["source"] == "DUI Defense"
    assert len(mapping["concepts"]) == 1
    assert mapping["concepts"][0]["label"] == "Criminal Law"


# --- RDF/Turtle tests ---


def test_rdf_turtle_has_valid_syntax():
    req = _make_request(options=_make_options(format="rdf_turtle"))
    data = generate_rdf_turtle(req)
    text = data.decode("utf-8")
    assert "@prefix folio:" in text
    assert "folio:RCriminalLaw" in text
    assert 'rdfs:label "Criminal Law"' in text
    assert 'dcterms:subject "DUI Defense"' in text


def test_rdf_turtle_escapes_quotes():
    concept = _make_concept(label='Law of "Torts"', definition=None)
    row = _make_row(selected_concepts=[concept])
    req = _make_request(rows=[row])
    data = generate_rdf_turtle(req)
    text = data.decode("utf-8")
    assert r'Law of \"Torts\"' in text


# --- JSON-LD tests ---


def test_json_ld_has_context_and_graph():
    req = _make_request(options=_make_options(format="json_ld"))
    data = generate_json_ld(req)
    parsed = json.loads(data)
    assert "@context" in parsed
    assert "@graph" in parsed
    assert "folio" in parsed["@context"]
    node = parsed["@graph"][0]
    assert node["@id"] == "folio:RCriminalLaw"
    assert node["rdfs:label"] == "Criminal Law"


def test_json_ld_includes_translations():
    req = _make_request(options=_make_options(format="json_ld"))
    data = generate_json_ld(req)
    parsed = json.loads(data)
    node = parsed["@graph"][0]
    assert "skos:altLabel" in node
    labels = {e["@language"]: e["@value"] for e in node["skos:altLabel"]}
    assert labels["es"] == "Derecho Penal"


# --- Markdown tests ---


def test_markdown_generates_table():
    req = _make_request(options=_make_options(format="markdown"))
    data = generate_markdown(req)
    text = data.decode("utf-8")
    assert "# FOLIO Mapping Results" in text
    assert "| Source | Label | IRI | Branch | Confidence | Notes |" in text
    assert "| DUI Defense | Criminal Law |" in text


# --- HTML tests ---


def test_html_generates_valid_document():
    req = _make_request(options=_make_options(format="html"))
    data = generate_html(req)
    text = data.decode("utf-8")
    assert "<!DOCTYPE html>" in text
    assert "<title>FOLIO Mapping Report</title>" in text
    assert "DUI Defense" in text
    assert "Criminal Law" in text


def test_html_confidence_color_classes():
    concepts = [
        _make_concept(score=95.0),
        _make_concept(iri_hash="R2", score=78.0),
        _make_concept(iri_hash="R3", score=63.0),
        _make_concept(iri_hash="R4", score=50.0),
        _make_concept(iri_hash="R5", score=30.0),
    ]
    rows = [_make_row(selected_concepts=[c]) for c in concepts]
    req = _make_request(rows=rows, options=_make_options(format="html"))
    data = generate_html(req)
    text = data.decode("utf-8")
    assert 'class="conf-90"' in text
    assert 'class="conf-75"' in text
    assert 'class="conf-60"' in text
    assert 'class="conf-45"' in text
    assert 'class="conf-low"' in text


# --- IRI format tests ---


def test_iri_format_full_url():
    req = _make_request(options=_make_options(iri_format="full_url"))
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    assert "https://folio.openlegalstandard.org/RCriminalLaw" in text


def test_iri_format_both():
    req = _make_request(options=_make_options(iri_format="both"))
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    assert "RCriminalLaw (https://folio.openlegalstandard.org/RCriminalLaw)" in text


# --- Column toggle tests ---


def test_exclude_confidence():
    req = _make_request(options=_make_options(include_confidence=False))
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    headers = next(reader)
    assert "Confidence" not in headers


def test_exclude_notes():
    req = _make_request(options=_make_options(include_notes=False))
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    headers = next(reader)
    assert "Notes" not in headers


# --- Preview endpoint tests ---


@pytest.mark.anyio
async def test_preview_endpoint_returns_rows(client: AsyncClient):
    req = _make_request()
    resp = await client.post("/api/export/preview", json=req.model_dump())
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["source"] == "DUI Defense"
    assert data[0]["label"] == "Criminal Law"


@pytest.mark.anyio
async def test_preview_limits_to_5_rows(client: AsyncClient):
    rows = [_make_row(item_index=i, source_text=f"Item {i}") for i in range(10)]
    req = _make_request(rows=rows)
    resp = await client.post("/api/export/preview", json=req.model_dump())
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) <= 5


# --- Generate endpoint tests ---


@pytest.mark.anyio
async def test_generate_csv_endpoint(client: AsyncClient):
    req = _make_request()
    resp = await client.post("/api/export/generate", json=req.model_dump())
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    assert "folio-mappings.csv" in resp.headers.get("content-disposition", "")


@pytest.mark.anyio
async def test_generate_unsupported_format(client: AsyncClient):
    req = _make_request(options=_make_options(format="invalid"))
    resp = await client.post("/api/export/generate", json=req.model_dump())
    assert resp.status_code == 400


# --- Translation endpoint test ---


@pytest.mark.anyio
async def test_translations_endpoint(client: AsyncClient):
    mock_folio = MagicMock()
    mock_class = MagicMock()
    mock_class.translations = {"es": "Derecho Penal"}
    mock_folio.__getitem__ = MagicMock(return_value=mock_class)

    with patch("app.services.export_service.get_folio", return_value=mock_folio):
        resp = await client.post(
            "/api/export/translations",
            json={"iri_hashes": ["RCriminalLaw"], "languages": ["es"]},
        )
    assert resp.status_code == 200
    data = resp.json()
    assert data["RCriminalLaw"]["es"] == "Derecho Penal"


# --- Export scope tests ---


def _mock_owl_class(
    iri_hash: str = "RCriminalLaw",
    label: str = "Criminal Law",
    definition: str | None = "Law dealing with crimes",
    sub_class_of: list[str] | None = None,
    parent_class_of: list[str] | None = None,
    alternative_labels: list[str] | None = None,
    examples: list[str] | None = None,
    translations: dict | None = None,
    see_also: list[str] | None = None,
    deprecated: bool = False,
) -> MagicMock:
    cls = MagicMock()
    cls.iri = f"https://folio.openlegalstandard.org/{iri_hash}"
    cls.label = label
    cls.definition = definition
    cls.sub_class_of = sub_class_of or ["http://www.w3.org/2002/07/owl#Thing"]
    cls.parent_class_of = parent_class_of
    cls.alternative_labels = alternative_labels or []
    cls.examples = examples or []
    cls.translations = translations or {}
    cls.see_also = see_also
    cls.editorial_note = None
    cls.history_note = None
    cls.deprecated = deprecated
    return cls


def _make_mock_folio(classes_map: dict[str, MagicMock]) -> MagicMock:
    """Create a mock FOLIO that can look up classes by iri_hash."""
    folio = MagicMock()
    folio.__getitem__ = MagicMock(side_effect=lambda h: classes_map.get(h))
    folio.classes = [classes_map[h] for h in classes_map]
    return folio


def test_enrich_concept_basic():
    """enrich_concept returns full metadata for a known class."""
    owl = _mock_owl_class(
        alternative_labels=["Penal Law"],
        examples=["Murder", "Theft"],
        translations={"es": "Derecho Penal"},
    )
    folio = _make_mock_folio({"RCriminalLaw": owl})

    with patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"):
        result = enrich_concept(
            folio, "RCriminalLaw",
            score=85.0, is_mapped=True,
            mapping_source_text="DUI Defense",
            relationship="direct",
        )

    assert result is not None
    assert result.label == "Criminal Law"
    assert result.iri_hash == "RCriminalLaw"
    assert result.alternative_labels == ["Penal Law"]
    assert result.examples == ["Murder", "Theft"]
    assert result.translations == {"es": "Derecho Penal"}
    assert result.is_mapped is True
    assert result.relationship == "direct"
    assert result.mapping_source_text == "DUI Defense"
    assert result.score == 85.0
    assert result.branch == "Area of Law"


def test_enrich_concept_returns_none_for_unknown():
    """enrich_concept returns None if iri_hash not found."""
    folio = _make_mock_folio({})

    with patch("app.services.export_scope.get_branch_for_class", return_value="Unknown"):
        result = enrich_concept(folio, "RDoesNotExist")

    assert result is None


def test_enrich_concept_hierarchy_path():
    """enrich_concept builds hierarchy path from parent chain."""
    parent = _mock_owl_class(
        iri_hash="RAreaOfLaw", label="Area of Law",
        sub_class_of=["http://www.w3.org/2002/07/owl#Thing"],
    )
    child = _mock_owl_class(
        iri_hash="RCriminalLaw", label="Criminal Law",
        sub_class_of=["https://folio.openlegalstandard.org/RAreaOfLaw"],
    )
    folio = _make_mock_folio({"RCriminalLaw": child, "RAreaOfLaw": parent})

    with patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"):
        result = enrich_concept(folio, "RCriminalLaw", score=85.0)

    assert result is not None
    assert result.hierarchy_path == ["Area of Law", "Criminal Law"]


def test_enrich_concept_deprecated_flag():
    """enrich_concept picks up the deprecated flag."""
    owl = _mock_owl_class(deprecated=True)
    folio = _make_mock_folio({"RCriminalLaw": owl})

    with patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"):
        result = enrich_concept(folio, "RCriminalLaw", score=0.0)

    assert result is not None
    assert result.deprecated is True


def test_expand_scope_mapped_only_enriches():
    """mapped_only scope enriches concepts with full metadata."""
    owl = _mock_owl_class(alternative_labels=["Penal Law"])
    folio = _make_mock_folio({"RCriminalLaw": owl})
    req = _make_request(options=_make_options(export_scope="mapped_only"))

    with (
        patch("app.services.export_scope.get_folio", return_value=folio),
        patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"),
    ):
        result = expand_scope(req)

    assert len(result.rows) == 1
    concept = result.rows[0].selected_concepts[0]
    assert concept.is_mapped is True
    assert concept.relationship == "direct"
    assert concept.alternative_labels == ["Penal Law"]


def test_expand_scope_mapped_with_related_adds_descendants():
    """mapped_with_related scope adds direct descendants of mapped concepts."""
    parent = _mock_owl_class(
        iri_hash="RAreaOfLaw", label="Area of Law",
        sub_class_of=["http://www.w3.org/2002/07/owl#Thing"],
        parent_class_of=[
            "https://folio.openlegalstandard.org/RCriminalLaw",
        ],
    )
    criminal = _mock_owl_class(
        iri_hash="RCriminalLaw", label="Criminal Law",
        sub_class_of=["https://folio.openlegalstandard.org/RAreaOfLaw"],
        parent_class_of=[
            "https://folio.openlegalstandard.org/RDUILaw",
            "https://folio.openlegalstandard.org/RTheftLaw",
        ],
    )
    dui = _mock_owl_class(
        iri_hash="RDUILaw", label="DUI Law",
        sub_class_of=["https://folio.openlegalstandard.org/RCriminalLaw"],
    )
    theft = _mock_owl_class(
        iri_hash="RTheftLaw", label="Theft Law",
        sub_class_of=["https://folio.openlegalstandard.org/RCriminalLaw"],
    )
    folio = _make_mock_folio({
        "RCriminalLaw": criminal, "RDUILaw": dui, "RTheftLaw": theft, "RAreaOfLaw": parent,
    })
    req = _make_request(options=_make_options(export_scope="mapped_with_related"))

    with (
        patch("app.services.export_scope.get_folio", return_value=folio),
        patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"),
    ):
        result = expand_scope(req)

    concepts = result.rows[0].selected_concepts
    hashes = [c.iri_hash for c in concepts]
    assert "RCriminalLaw" in hashes  # mapped concept
    assert "RDUILaw" in hashes  # direct child
    assert "RTheftLaw" in hashes  # direct child

    direct = next(c for c in concepts if c.iri_hash == "RCriminalLaw")
    assert direct.is_mapped is True
    assert direct.relationship == "direct"

    child = next(c for c in concepts if c.iri_hash == "RDUILaw")
    assert child.is_mapped is False
    assert child.relationship == "child"


def test_expand_scope_mapped_with_related_adds_ancestors():
    """mapped_with_related scope adds ancestors up to branch root."""
    grandparent = _mock_owl_class(
        iri_hash="RAreaOfLaw", label="Area of Law",
        sub_class_of=["http://www.w3.org/2002/07/owl#Thing"],
        parent_class_of=["https://folio.openlegalstandard.org/RCriminalLaw"],
    )
    parent_cls = _mock_owl_class(
        iri_hash="RCriminalLaw", label="Criminal Law",
        sub_class_of=["https://folio.openlegalstandard.org/RAreaOfLaw"],
        parent_class_of=["https://folio.openlegalstandard.org/RDUILaw"],
    )
    child = _mock_owl_class(
        iri_hash="RDUILaw", label="DUI Law",
        sub_class_of=["https://folio.openlegalstandard.org/RCriminalLaw"],
    )
    folio = _make_mock_folio({
        "RDUILaw": child, "RCriminalLaw": parent_cls, "RAreaOfLaw": grandparent,
    })
    concept = _make_concept(iri_hash="RDUILaw", label="DUI Law")
    row = _make_row(selected_concepts=[concept])
    req = _make_request(rows=[row], options=_make_options(export_scope="mapped_with_related"))

    with (
        patch("app.services.export_scope.get_folio", return_value=folio),
        patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"),
    ):
        result = expand_scope(req)

    concepts = result.rows[0].selected_concepts
    hashes = [c.iri_hash for c in concepts]
    # Should have the mapped concept + ancestor (Criminal Law)
    # RAreaOfLaw's parent is owl:Thing so ancestor walk stops before it
    assert "RDUILaw" in hashes
    assert "RCriminalLaw" in hashes
    ancestor = next(c for c in concepts if c.iri_hash == "RCriminalLaw")
    assert ancestor.relationship == "ancestor"
    assert ancestor.is_mapped is False


def test_expand_scope_mapped_with_related_deduplicates():
    """mapped_with_related scope deduplicates by iri_hash within a row."""
    parent = _mock_owl_class(
        iri_hash="RAreaOfLaw", label="Area of Law",
        sub_class_of=["http://www.w3.org/2002/07/owl#Thing"],
        parent_class_of=["https://folio.openlegalstandard.org/RCriminalLaw"],
    )
    criminal = _mock_owl_class(
        iri_hash="RCriminalLaw", label="Criminal Law",
        sub_class_of=["https://folio.openlegalstandard.org/RAreaOfLaw"],
    )
    folio = _make_mock_folio({"RCriminalLaw": criminal, "RAreaOfLaw": parent})
    req = _make_request(options=_make_options(export_scope="mapped_with_related"))

    with (
        patch("app.services.export_scope.get_folio", return_value=folio),
        patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"),
    ):
        result = expand_scope(req)

    hashes = [c.iri_hash for c in result.rows[0].selected_concepts]
    assert hashes.count("RCriminalLaw") == 1


def test_expand_scope_full_ontology():
    """full_ontology scope includes all classes with mapped flagging."""
    criminal = _mock_owl_class(iri_hash="RCriminalLaw", label="Criminal Law")
    civil = _mock_owl_class(iri_hash="RCivilLaw", label="Civil Law")
    folio = _make_mock_folio({"RCriminalLaw": criminal, "RCivilLaw": civil})
    req = _make_request(options=_make_options(export_scope="full_ontology"))

    with (
        patch("app.services.export_scope.get_folio", return_value=folio),
        patch("app.services.export_scope.get_branch_for_class", return_value="Area of Law"),
    ):
        result = expand_scope(req)

    # Should have rows for all classes
    all_hashes = set()
    for row in result.rows:
        for c in row.selected_concepts:
            all_hashes.add(c.iri_hash)
    assert "RCriminalLaw" in all_hashes
    assert "RCivilLaw" in all_hashes

    # RCriminalLaw was in the original request, so it should be mapped
    for row in result.rows:
        for c in row.selected_concepts:
            if c.iri_hash == "RCriminalLaw":
                assert c.is_mapped is True
                assert c.relationship == "direct"
            elif c.iri_hash == "RCivilLaw":
                assert c.is_mapped is False
                assert c.relationship == "ontology"


# --- Format tests with scope columns ---


def _make_enriched_concept(**overrides) -> ExportConcept:
    """Create an ExportConcept with all metadata fields populated."""
    defaults = {
        "label": "Criminal Law",
        "iri": "https://folio.openlegalstandard.org/RCriminalLaw",
        "iri_hash": "RCriminalLaw",
        "branch": "Area of Law",
        "score": 85.0,
        "definition": "Law dealing with crimes and punishments",
        "translations": {"es": "Derecho Penal"},
        "alternative_labels": ["Penal Law"],
        "examples": ["Murder", "Theft"],
        "hierarchy_path": ["Area of Law", "Criminal Law"],
        "parent_iri_hash": "RAreaOfLaw",
        "see_also": ["RCivilLaw"],
        "notes": None,
        "deprecated": False,
        "is_mapped": True,
        "mapping_source_text": "DUI Defense",
        "relationship": "direct",
    }
    defaults.update(overrides)
    return ExportConcept(**defaults)


def _make_scope_request(scope: str = "mapped_only", **overrides) -> ExportRequest:
    """Create a request with enriched concepts for testing format output."""
    concept = _make_enriched_concept()
    row = ExportRow(
        item_index=0,
        source_text="DUI Defense",
        ancestry=["Criminal Law"],
        selected_concepts=[concept],
        note="Reviewed",
        status="completed",
    )
    defaults = {
        "rows": [row],
        "options": _make_options(export_scope=scope),
        "source_file": "test.txt",
        "session_created": None,
    }
    defaults.update(overrides)
    return ExportRequest(**defaults)


def test_csv_metadata_columns_mapped_only():
    """CSV includes metadata columns (Definition, Alt Labels, etc.) for mapped_only."""
    req = _make_scope_request("mapped_only")
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    headers = rows[0]
    assert "Definition" in headers
    assert "Alt Labels" in headers
    assert "Examples" in headers
    assert "Hierarchy" in headers
    # mapped_only should NOT have scope columns
    assert "Mapped" not in headers
    assert "Relationship" not in headers
    # Check data values
    def_idx = headers.index("Definition")
    assert rows[1][def_idx] == "Law dealing with crimes and punishments"
    alt_idx = headers.index("Alt Labels")
    assert rows[1][alt_idx] == "Penal Law"
    hier_idx = headers.index("Hierarchy")
    assert rows[1][hier_idx] == "Area of Law > Criminal Law"


def test_csv_scope_columns_mapped_with_related():
    """CSV includes Mapped/Relationship/Source Text columns for non-mapped_only scopes."""
    req = _make_scope_request("mapped_with_related")
    data = generate_csv(req)
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    headers = rows[0]
    assert "Mapped" in headers
    assert "Relationship" in headers
    assert "Source Text" in headers
    mapped_idx = headers.index("Mapped")
    assert rows[1][mapped_idx] == "Yes"
    rel_idx = headers.index("Relationship")
    assert rows[1][rel_idx] == "direct"


def test_json_includes_all_metadata():
    """JSON output includes all enriched metadata fields."""
    req = _make_scope_request("mapped_only")
    data = generate_json(req)
    parsed = json.loads(data)
    concept = parsed["mappings"][0]["concepts"][0]
    assert concept["alternative_labels"] == ["Penal Law"]
    assert concept["examples"] == ["Murder", "Theft"]
    assert concept["hierarchy_path"] == ["Area of Law", "Criminal Law"]
    assert concept["parent_iri_hash"] == "RAreaOfLaw"
    assert concept["see_also"] == ["RCivilLaw"]
    assert parsed["export_scope"] == "mapped_only"


def test_json_scope_columns_for_related():
    """JSON includes is_mapped/relationship for non-mapped_only scope."""
    req = _make_scope_request("mapped_with_related")
    data = generate_json(req)
    parsed = json.loads(data)
    concept = parsed["mappings"][0]["concepts"][0]
    assert concept["is_mapped"] is True
    assert concept["relationship"] == "direct"
    assert concept["mapping_source_text"] == "DUI Defense"


def test_json_no_scope_columns_for_mapped_only():
    """JSON omits is_mapped/relationship for mapped_only scope."""
    req = _make_scope_request("mapped_only")
    data = generate_json(req)
    parsed = json.loads(data)
    concept = parsed["mappings"][0]["concepts"][0]
    assert "is_mapped" not in concept
    assert "relationship" not in concept


def test_rdf_turtle_includes_metadata():
    """RDF/Turtle includes altLabel, example, broader, seeAlso."""
    req = _make_scope_request("mapped_only", options=_make_options(format="rdf_turtle"))
    data = generate_rdf_turtle(req)
    text = data.decode("utf-8")
    assert 'skos:altLabel "Penal Law"' in text
    assert 'skos:example "Murder"' in text
    assert "skos:broader folio:RAreaOfLaw" in text
    assert "rdfs:seeAlso folio:RCivilLaw" in text


def test_rdf_turtle_scope_annotations():
    """RDF/Turtle includes isMapped annotation for non-mapped_only."""
    req = _make_scope_request(
        "mapped_with_related",
        options=_make_options(format="rdf_turtle", export_scope="mapped_with_related"),
    )
    data = generate_rdf_turtle(req)
    text = data.decode("utf-8")
    assert 'folio-mapper:isMapped "true"' in text
    assert 'folio-mapper:mappedFrom "DUI Defense"' in text


def test_rdf_turtle_no_scope_annotations_mapped_only():
    """RDF/Turtle omits scope annotations for mapped_only."""
    req = _make_scope_request("mapped_only", options=_make_options(format="rdf_turtle"))
    data = generate_rdf_turtle(req)
    text = data.decode("utf-8")
    assert "folio-mapper:isMapped" not in text


def test_json_ld_includes_metadata():
    """JSON-LD includes broader, seeAlso, altLabel for enriched concepts."""
    req = _make_scope_request("mapped_only", options=_make_options(format="json_ld"))
    data = generate_json_ld(req)
    parsed = json.loads(data)
    node = parsed["@graph"][0]
    assert node["skos:broader"] == {"@id": "folio:RAreaOfLaw"}
    assert {"@id": "folio:RCivilLaw"} in node["rdfs:seeAlso"]
    # Alt labels include translations + alternative_labels
    alt_values = [a.get("@value") for a in node["skos:altLabel"]]
    assert "Penal Law" in alt_values
    assert "Derecho Penal" in alt_values


def test_json_ld_scope_annotations():
    """JSON-LD includes folio-mapper:isMapped for non-mapped_only."""
    req = _make_scope_request(
        "full_ontology",
        options=_make_options(format="json_ld", export_scope="full_ontology"),
    )
    data = generate_json_ld(req)
    parsed = json.loads(data)
    assert "folio-mapper" in parsed["@context"]
    node = parsed["@graph"][0]
    assert node["folio-mapper:isMapped"] is True


def test_html_mapped_row_border():
    """HTML adds mapped-row class to mapped rows for non-mapped_only scopes."""
    concept_mapped = _make_enriched_concept(is_mapped=True, relationship="direct")
    concept_unmapped = _make_enriched_concept(
        iri_hash="RCivilLaw", label="Civil Law",
        is_mapped=False, relationship="sibling",
    )
    rows = [
        ExportRow(
            item_index=0, source_text="DUI Defense", ancestry=[],
            selected_concepts=[concept_mapped], note=None, status="completed",
        ),
        ExportRow(
            item_index=1, source_text="", ancestry=[],
            selected_concepts=[concept_unmapped], note=None, status="completed",
        ),
    ]
    req = ExportRequest(
        rows=rows,
        options=_make_options(format="html", export_scope="mapped_with_related"),
        source_file="test.txt",
        session_created=None,
    )
    data = generate_html(req)
    text = data.decode("utf-8")
    assert 'class="mapped-row"' in text
    assert ".mapped-row { border-left: 3px solid #22c55e; }" in text


def test_excel_metadata_columns():
    """Excel includes metadata columns."""
    req = _make_scope_request("mapped_only")
    data = generate_excel(req)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [ws.cell(1, i + 1).value for i in range(ws.max_column)]
    assert "Definition" in headers
    assert "Alt Labels" in headers
    assert "Examples" in headers
    assert "Hierarchy" in headers
